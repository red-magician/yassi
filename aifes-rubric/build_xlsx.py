# -*- coding: utf-8 -*-
"""rubric_data.py から審査用ワークブックを生成する。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

from rubric_data import CRITERIA, JUDGES, FLAGS, BANDS, TIEBREAK, META, WEIGHTS, STEP

RAW_TOTAL = sum(c["raw"] for c in CRITERIA)

FONT = "Arial"
NAVY = "1F3864"
GOLD = "BF8F00"
LGRAY = "F2F2F2"
LBLUE = "DDEBF7"
LYELL = "FFF2CC"
LGREEN = "E2EFDA"

H1 = Font(name=FONT, size=16, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
SMALL = Font(name=FONT, size=9, color="595959")
INPUT_F = Font(name=FONT, size=10, bold=True, color="0000FF")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_GRAY = PatternFill("solid", fgColor=LGRAY)
FILL_BLUE = PatternFill("solid", fgColor=LBLUE)
FILL_YELL = PatternFill("solid", fgColor=LYELL)
FILL_GREEN = PatternFill("solid", fgColor=LGREEN)

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(horizontal="center", vertical="center")

JUDGE_BY_NO = {j["no"]: j for j in JUDGES}
LEVEL_LABELS = [5, 4, 3, 2, 1]


def banner(ws, text, width, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, text)
    c.font, c.fill, c.alignment = H1, FILL_NAVY, Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        c = ws.cell(2, 1, sub)
        c.font, c.fill = SMALL, FILL_GRAY
        c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[2].height = 26


def head_row(ws, row, values, widths=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font, c.fill, c.alignment, c.border = HDR, FILL_NAVY, WRAP_C, BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


wb = openpyxl.Workbook()

# =============================================================== 00 はじめに
ws = wb.active
ws.title = "00_はじめに"
banner(ws, META["title"], 4)
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 96
ws.column_dimensions["D"].width = 2

rows = [
    ("目的", "グランドフィナーレの予備審査を、12名の審査員それぞれに個別に依頼することなくAIで一次実施するための共通ルーブリック。"
             "審査員12名のコメントを6観点に集約しており、1回の採点で全員の観点が反映される。"),
    ("元データ", META["source"] + "（本ブックの 07_元データ_審査員コメント シートに全文を収録）"),
    ("集約方針", META["policy"]),
    ("配点の出し方", "各観点に紐づく審査員の言及を数え、グランドフィナーレコメント由来を1.0、審査コメント由来を0.5として合計し、"
                  "100点に正規化したうえで、扱いやすさのため10点刻みに丸めている。02_審査員対応表 シートで、どの審査員のどの発言がどの観点の何点分になっているかを1件ずつ追える。"),
    ("採点の手順", "① 03_採点シート に応募部門を1行ずつ並べる  →  ② 各観点を 01_ルーブリック のレベル記述に照らして 1〜5 で入力  →  "
                 "③ 前提フラグ F1〜F4 に該当すれば ○ を入力（該当観点に自動で上限がかかる）  →  ④ 加重得点・総合点・判定・順位は自動計算。"),
    ("応募が5〜6件のとき", "04_役員予備審査投票 シートで、役員4名がAIの評価を参考に上位4部へ順位をつける。"
                       "1位10点・2位8点・3位7点・4位6点の配点と「各役員の1位は無条件で決勝進出」のルールで、決勝4部を自動判定する。"
                       "応募が4件以下ならこのシートは使わず、全件がそのまま決勝進出する。"),
    ("AIで実施する場合", "06_AI審査プロンプト シートの本文をそのままAIに渡し、応募資料を1件ずつ入力する。JSONで返させ、03_採点シート に貼り付ける。"),
    ("予備審査の位置づけ", "本ルーブリックは足切りと順位づけの一次スクリーニングに用いる。最終決定は審査員による人手審査で行う。"
                      "判定 S・A は人手審査へ自動送付、B は事務局が拾い上げを検討、C は見送りを基本とする。"),
    ("注意", "審査員本人の発言を観点に集約する過程で、個別の言い回しは要約されている。判断に迷う応募は、02_審査員対応表 から"
           "該当審査員の原文に戻って確認すること。"),
]
r = 3
for k, v in rows:
    ws.cell(r, 2, k).font = H2
    ws.cell(r, 2).fill = FILL_BLUE
    ws.cell(r, 2).alignment = WRAP
    ws.cell(r, 2).border = BOX
    c = ws.cell(r, 3, v)
    c.font, c.alignment, c.border = BODY, WRAP, BOX
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
ws.cell(r, 2, "観点と配点").font = Font(name=FONT, size=12, bold=True, color=GOLD)
r += 1
head_row(ws, r, ["", "観点", "配点 / 100"])
ws.cell(r, 2).value = "観点"
ws.cell(r, 3).value = "配点（100点満点） — 内訳は 01_ルーブリック 参照"
r += 1
for c_ in CRITERIA:
    a = ws.cell(r, 2, f"{c_['id']}  {c_['name']}")
    a.font, a.border, a.alignment = BODY_B, BOX, WRAP
    b = ws.cell(r, 3, f"{c_['weight']} 点　（言及重み {c_['raw']:.1f} / {RAW_TOTAL:.1f} ＝ {c_['exact']:.1f}% → {STEP}点刻みに丸めて {c_['weight']}）"
       f"　問い：{c_['question']}")
    b.font, b.border, b.alignment = BODY, BOX, WRAP
    ws.row_dimensions[r].height = 30
    r += 1
a = ws.cell(r, 2, "合計")
a.font, a.fill, a.border = BODY_B, FILL_YELL, BOX
b = ws.cell(r, 3, f"{sum(WEIGHTS)} 点")
b.font, b.fill, b.border = BODY_B, FILL_YELL, BOX

# =============================================================== 01 ルーブリック
ws = wb.create_sheet("01_ルーブリック")
banner(ws, "6観点 × 5段階 評価基準", 8,
       "レベルは 5＝満点相当、1＝最低。加重得点 ＝ レベル ÷ 5 × 配点。各観点の根拠となった審査員は右端を参照（◎＝グランドフィナーレコメント由来、○＝審査コメント由来）。")
widths = [10, 26, 9, 7, 70, 34, 2, 2]
head_row(ws, 3, ["観点ID", "観点名", "配点", "レベル", "レベルの記述（この状態なら該当）", "この観点の根拠となった審査員"], widths)
ws.row_dimensions[3].height = 32

r = 4
for c_ in CRITERIA:
    top = r
    for li, lvl in enumerate(LEVEL_LABELS):
        cell_lv = ws.cell(r, 4, lvl)
        cell_lv.font, cell_lv.alignment, cell_lv.border = BODY_B, CENTER, BOX
        cell_lv.fill = PatternFill("solid", fgColor=["C6E0B4", "E2EFDA", "FFF2CC", "FCE4D6", "F8CBAD"][li])
        d = ws.cell(r, 5, c_["levels"][li])
        d.font, d.alignment, d.border = BODY, WRAP, BOX
        ws.row_dimensions[r].height = 58
        r += 1
    bot = r - 1
    for col, val, fnt in ((1, c_["id"], BODY_B), (2, c_["name"], BODY_B), (3, c_["weight"], Font(name=FONT, size=14, bold=True, color=GOLD))):
        ws.merge_cells(start_row=top, start_column=col, end_row=bot, end_column=col)
        cc = ws.cell(top, col, val)
        cc.font, cc.alignment, cc.border = fnt, WRAP_C, BOX
        cc.fill = FILL_BLUE if col != 3 else FILL_YELL
    ws.cell(top, 1).comment = Comment(f"問い: {c_['question']}\n\n{c_['definition']}", "AIFES事務局", width=380, height=200)

    ws.merge_cells(start_row=top, start_column=6, end_row=bot, end_column=6)
    txt = "\n".join(
        f"{m} {JUDGE_BY_NO[jn]['name']}（{JUDGE_BY_NO[jn]['title']}）\n　　{note}"
        for jn, m, note in c_["sources"])
    sc = ws.cell(top, 6, txt)
    sc.font, sc.alignment, sc.border = Font(name=FONT, size=8), WRAP, BOX

    # 観点の定義を1行はさむ
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    dcell = ws.cell(r, 1, f"【{c_['id']} の問い】{c_['question']}　／　{c_['definition']}")
    dcell.font, dcell.fill, dcell.alignment, dcell.border = SMALL, FILL_GRAY, WRAP, BOX
    ws.row_dimensions[r].height = 30
    r += 2

ws.freeze_panes = "D4"

# =============================================================== 02 審査員対応表
ws = wb.create_sheet("02_審査員対応表")
banner(ws, "審査員12名 × 6観点 カバレッジ対応表", 10,
       "◎＝グランドフィナーレコメントが直接この観点の根拠になっている　／　○＝審査コメント（ステージ用）が補助的に反映されている。"
       "12名全員が最低1観点に◎で反映されている。")
widths = [5, 16, 26] + [11] * 6 + [2]
head_row(ws, 3, ["No", "審査員", "役職"] + [f"{c_['id']}\n{c_['short']}\n({c_['weight']}点)" for c_ in CRITERIA], widths)
ws.row_dimensions[3].height = 46

r = 4
for j in JUDGES:
    ws.cell(r, 1, j["no"]).font = BODY
    ws.cell(r, 2, j["name"]).font = BODY_B
    ws.cell(r, 3, j["title"]).font = BODY
    for col in (1, 2, 3):
        ws.cell(r, col).border, ws.cell(r, col).alignment = BOX, WRAP
    for ci, c_ in enumerate(CRITERIA):
        marks = [(m, note) for jn, m, note in c_["sources"] if jn == j["no"]]
        cell = ws.cell(r, 4 + ci)
        cell.border, cell.alignment, cell.font = BOX, CENTER, Font(name=FONT, size=12, bold=True)
        if marks:
            mk = "◎" if any(m == "◎" for m, _ in marks) else "○"
            cell.value = mk
            cell.fill = FILL_GREEN if mk == "◎" else FILL_YELL
            cell.comment = Comment("\n".join(f"{m}  {n}" for m, n in marks), "AIFES事務局", width=340, height=120)
    ws.row_dimensions[r].height = 26
    r += 1

# 合計行
ws.cell(r, 3, "◎ の数 / ○ の数").font = BODY_B
ws.cell(r, 3).fill, ws.cell(r, 3).border, ws.cell(r, 3).alignment = FILL_YELL, BOX, CENTER
for ci, c_ in enumerate(CRITERIA):
    n_d = sum(1 for _, m, _ in c_["sources"] if m == "◎")
    n_o = sum(1 for _, m, _ in c_["sources"] if m == "○")
    cell = ws.cell(r, 4 + ci, f"{n_d} / {n_o}")
    cell.font, cell.fill, cell.border, cell.alignment = BODY_B, FILL_YELL, BOX, CENTER
r += 1
ws.cell(r, 3, "言及重み（◎×1.0＋○×0.5）").font = BODY_B
ws.cell(r, 3).fill, ws.cell(r, 3).border, ws.cell(r, 3).alignment = FILL_GRAY, BOX, CENTER
for ci, c_ in enumerate(CRITERIA):
    cell = ws.cell(r, 4 + ci, c_["raw"])
    cell.font, cell.fill, cell.border, cell.alignment = BODY, FILL_GRAY, BOX, CENTER
r += 1
ws.cell(r, 3, "→ 配点（100点に正規化 → 10点刻み）").font = BODY_B
ws.cell(r, 3).fill, ws.cell(r, 3).border, ws.cell(r, 3).alignment = FILL_GOLD, BOX, CENTER
ws.cell(r, 3).font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
for ci, c_ in enumerate(CRITERIA):
    cell = ws.cell(r, 4 + ci, c_["weight"])
    cell.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    cell.fill, cell.border, cell.alignment = FILL_GOLD, BOX, CENTER

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
note = ws.cell(r, 1, "各審査員の発言原文は 07_元データ_審査員コメント シートに全文収録。セルのコメント（赤い三角）に、その審査員のどの発言がこの観点に効いているかを記載している。")
note.font, note.fill, note.alignment = SMALL, FILL_GRAY, WRAP
ws.freeze_panes = "D4"

# =============================================================== 03 採点シート
ws = wb.create_sheet("03_採点シート")
N_ROWS = 30
HDR_ROW = 4
W_ROW = 5           # 配点行
EX_ROW = 6          # 記入例
FIRST = 7
LAST = FIRST + N_ROWS - 1

banner(ws, "採点シート", 25,
       "青字のセルだけ入力する（観点レベル 1〜5、前提フラグ ○）。加重得点・総合点・判定・順位は自動計算。"
       "配点は5行目に置いてあり、変更すれば全行に反映される。")

cols = (["No", "部門名", "部門長"]
        + [f"{c_['id']}\n{c_['short']}\nレベル1-5" for c_ in CRITERIA]
        + [f"{c_['id']}\n加重" for c_ in CRITERIA]
        + ["総合点\n(100)", "判定", "順位"]
        + [f"{f['id']}\n{f['name']}" for f in FLAGS]
        + ["上限チェック", "AI所見（要約）", "人手審査へ"])
widths = [5, 24, 14] + [10] * 6 + [8] * 6 + [9, 7, 7] + [11] * 4 + [22, 46, 11]
head_row(ws, HDR_ROW, cols, widths)
ws.row_dimensions[HDR_ROW].height = 48

L1, L2 = 4, 9          # レベル入力 D..I
W1, W2 = 10, 15        # 加重     J..O
TOT, JDG, RNK = 16, 17, 18   # P, Q, R
FL1 = 19                     # S..V
CHK, MEMO, HUM = 23, 24, 25  # W, X, Y

# 配点行
ws.cell(W_ROW, 3, "配点 →").font = BODY_B
ws.cell(W_ROW, 3).alignment = Alignment(horizontal="right")
for i, c_ in enumerate(CRITERIA):
    cell = ws.cell(W_ROW, L1 + i, c_["weight"])
    cell.font = Font(name=FONT, size=11, bold=True, color="0000FF")
    cell.fill, cell.border, cell.alignment = FILL_YELL, BOX, CENTER
ws.cell(W_ROW, TOT, f"={'+'.join(get_column_letter(L1+i)+str(W_ROW) for i in range(6))}")
ws.cell(W_ROW, TOT).font, ws.cell(W_ROW, TOT).fill, ws.cell(W_ROW, TOT).border = BODY_B, FILL_YELL, BOX
ws.cell(W_ROW, MEMO, "↑ 配点（合計100）。ここを直せば全行の加重得点が変わる。").font = SMALL

WL = get_column_letter(L1)
example = dict(no=0, dept="（記入例）AIX事業本部", head="山田 太郎",
               levels=[4, 5, 3, 4, 4, 5], flags=["", "", "", ""],
               memo="全員参加率92%を数値提示。請求書突合を廃止。他部門展開は2部門で実績あり、全社資産化は未着手。",
               human="要")

for r in range(EX_ROW, LAST + 1):
    is_ex = (r == EX_ROW)
    ws.cell(r, 1, "例" if is_ex else r - EX_ROW).font = BODY
    ws.cell(r, 1).alignment = CENTER
    if is_ex:
        ws.cell(r, 2, example["dept"])
        ws.cell(r, 3, example["head"])
        for i, v in enumerate(example["levels"]):
            ws.cell(r, L1 + i, v)
        ws.cell(r, MEMO, example["memo"])
        ws.cell(r, HUM, example["human"])

    for col in range(1, HUM + 1):
        cell = ws.cell(r, col)
        cell.border = BOX
        cell.font = BODY
        cell.alignment = WRAP if col in (2, 3, MEMO) else CENTER
        if is_ex:
            cell.fill = FILL_GRAY

    # 入力セル（青字）
    for col in list(range(L1, L2 + 1)) + list(range(FL1, FL1 + 4)) + [MEMO, HUM, 2, 3]:
        ws.cell(r, col).font = INPUT_F if not is_ex else Font(name=FONT, size=10, italic=True, color="808080")

    # 加重得点
    for i in range(6):
        lv = f"{get_column_letter(L1+i)}{r}"
        wt = f"${get_column_letter(L1+i)}${W_ROW}"
        c = ws.cell(r, W1 + i, f'=IF({lv}="","",ROUND({lv}/5*{wt},1))')
        c.font, c.alignment = BODY, CENTER
        c.number_format = "0.#"

    wrange = f"{get_column_letter(W1)}{r}:{get_column_letter(W2)}{r}"
    ws.cell(r, TOT, f'=IF(COUNT({get_column_letter(L1)}{r}:{get_column_letter(L2)}{r})<6,"",ROUND(SUM({wrange}),1))')
    ws.cell(r, TOT).font = Font(name=FONT, size=11, bold=True)
    ws.cell(r, TOT).number_format = "0.#"

    t = f"{get_column_letter(TOT)}{r}"
    band_f = f'=IF({t}="","",'
    for cut, lab, _ in BANDS[:-1]:
        band_f += f'IF({t}>={cut},"{lab}",'
    band_f += f'"{BANDS[-1][1]}"' + ")" * len(BANDS[:-1]) + ")"
    ws.cell(r, JDG, band_f)
    ws.cell(r, JDG).font = Font(name=FONT, size=11, bold=True)

    # 記入例行は順位の集計対象外なので RANK を入れない（範囲外参照でエラーになる）
    if is_ex:
        ws.cell(r, RNK, "—")
    else:
        ws.cell(r, RNK, f'=IF({t}="","",RANK({t},${get_column_letter(TOT)}${FIRST}:${get_column_letter(TOT)}${LAST}))')

    # 上限チェック
    S = lambda k: f"{get_column_letter(FL1+k)}{r}"
    C = lambda cid: f"{get_column_letter(L1 + [x['id'] for x in CRITERIA].index(cid))}{r}"
    conds = []
    for k, f in enumerate(FLAGS):
        sub = " ,".join(f"{C(cid)}>{cap}" for cid, cap in f["cap"].items())
        inner = f"OR({sub})" if len(f["cap"]) > 1 else sub
        conds.append((f["id"], f'AND({S(k)}<>"",{inner})'))
    cnt = "+".join(f"IF({c},1,0)" for _, c in conds)
    detail = "&".join(f'IF({c},"{fid} ","")' for fid, c in conds)
    ws.cell(r, CHK, f'=IF(COUNT({get_column_letter(L1)}{r}:{get_column_letter(L2)}{r})=0,"",IF({cnt}=0,"OK","上限超過: "&{detail}))')
    ws.cell(r, CHK).font = BODY
    ws.cell(r, CHK).alignment = WRAP

ws.row_dimensions[EX_ROW].height = 34

# 入力規則
dv_lv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True,
                       error="観点レベルは 1〜5 の整数で入力してください。", errorTitle="入力値エラー")
ws.add_data_validation(dv_lv)
dv_lv.add(f"{get_column_letter(L1)}{FIRST}:{get_column_letter(L2)}{LAST}")

dv_fl = DataValidation(type="list", formula1='"○"', allow_blank=True,
                       error="該当する場合のみ ○ を入力してください。", errorTitle="入力値エラー")
ws.add_data_validation(dv_fl)
dv_fl.add(f"{get_column_letter(FL1)}{FIRST}:{get_column_letter(FL1+3)}{LAST}")

dv_hu = DataValidation(type="list", formula1='"要,不要"', allow_blank=True)
ws.add_data_validation(dv_hu)
dv_hu.add(f"{get_column_letter(HUM)}{FIRST}:{get_column_letter(HUM)}{LAST}")

# 凡例・集計
lr = LAST + 2
ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=9)
lg = ws.cell(lr, 1, "【入力するセル】青字の列のみ： 部門名・部門長 ／ D〜I（観点レベル 1〜5） ／ S〜V（前提フラグ、該当時に ○） ／ X（AI所見） ／ Y（人手審査へ 要/不要）。"
                    "　黒字の列（J〜R, W）は自動計算のため触らない。6行目は記入例なので集計・順位から除外してある。")
lg.font, lg.fill, lg.alignment = SMALL, FILL_YELL, WRAP
ws.row_dimensions[lr].height = 30

lr += 2
ws.cell(lr, 2, "集計").font = H2
for i, (lab, formula) in enumerate([
    ("応募件数", f'=COUNT(${get_column_letter(TOT)}${FIRST}:${get_column_letter(TOT)}${LAST})'),
    ("平均点", f'=IFERROR(ROUND(AVERAGE(${get_column_letter(TOT)}${FIRST}:${get_column_letter(TOT)}${LAST}),1),"")'),
    ("最高点", f'=IFERROR(MAX(${get_column_letter(TOT)}${FIRST}:${get_column_letter(TOT)}${LAST}),"")'),
    ("S 判定", f'=COUNTIF(${get_column_letter(JDG)}${FIRST}:${get_column_letter(JDG)}${LAST},"S")'),
    ("A 判定", f'=COUNTIF(${get_column_letter(JDG)}${FIRST}:${get_column_letter(JDG)}${LAST},"A")'),
    ("B 判定", f'=COUNTIF(${get_column_letter(JDG)}${FIRST}:${get_column_letter(JDG)}${LAST},"B")'),
    ("C 判定", f'=COUNTIF(${get_column_letter(JDG)}${FIRST}:${get_column_letter(JDG)}${LAST},"C")'),
    ("上限超過あり", f'=COUNTIF(${get_column_letter(CHK)}${FIRST}:${get_column_letter(CHK)}${LAST},"上限超過*")'),
], start=1):
    a = ws.cell(lr + i, 2, lab)
    a.font, a.border, a.fill = BODY_B, BOX, FILL_BLUE
    b = ws.cell(lr + i, 3, formula)
    b.font, b.border = BODY_B, BOX

ws.freeze_panes = f"D{FIRST}"

# =============================================================== 04 役員予備審査投票
# 応募が5〜6件のときだけ使う。4件以下なら全件がそのまま決勝進出のため不要。
# 01_ルーブリック（AI）の評価を参考に、役員4名が上位4部へ順位をつけ、
# 1位10点/2位8点/3位7点/4位6点＋「各役員の1位は無条件で決勝進出」の
# 1位保護ルールで決勝4部を決める（審査設計ボードで確定した方式）。
ws = wb.create_sheet("04_役員予備審査投票")
banner(ws, "役員予備審査 投票集計（応募5〜6件のときだけ使用）", 10,
       "応募4件以下なら全件がそのまま決勝進出。このシートは使わない。"
       "5〜6件のときだけ、役員4名がAIの評価（③列C・D＝03_採点シートより自動参照）を参考に"
       "上位4部へ順位をつけ、下の配点で決勝進出4部を自動判定する。")
for col, w in zip("ABCDEFGHIJ", [5, 26, 13, 10, 13, 13, 13, 13, 12, 10, 12]):
    ws.column_dimensions[col].width = w

APP_HEAD, APP_FIRST, APP_LAST = 5, 6, 11        # ①応募一覧（最大6件）
VOTE_HEAD, VOTE_FIRST, VOTE_LAST = 14, 15, 18   # ②役員4名の投票
SC_HEAD, SC_FIRST, SC_LAST = 21, 22, 27         # ③得点集計・決勝進出判定

app_rng = f"$B${APP_FIRST}:$B${APP_LAST}"
cond = (f'=IF(COUNTA({app_rng})=0,"部門名をまず①に入力してください。",'
        f'IF(COUNTA({app_rng})<=4,'
        f'"★ 応募"&COUNTA({app_rng})&"件（4件以下）：このシートは不要です。全件がそのまま決勝進出。②③は空欄のままで構いません。",'
        f'"応募"&COUNTA({app_rng})&"件：②で役員4名が投票すると、③で決勝4部を自動判定します。"))')
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)
cc = ws.cell(3, 2, cond)
cc.font = Font(name=FONT, size=11, bold=True, color=NAVY)
cc.fill, cc.alignment = FILL_YELL, Alignment(vertical="center", wrap_text=True, indent=1)
ws.row_dimensions[3].height = 30

# ---- ① 応募一覧 --------------------------------------------------------
ws.cell(APP_HEAD, 2, "① 応募一覧（部門名を入力。②のドロップダウンと③の集計の元になる）").font = H2
head_row(ws, APP_HEAD + 1, ["No", "部門名", "AI総合点（参考）", "AI判定（参考）"])
for i in range(APP_FIRST, APP_LAST + 1):
    ws.cell(i, 1, i - APP_FIRST + 1).font, ws.cell(i, 1).alignment = BODY, CENTER
    ws.cell(i, 2).font = INPUT_F
    ai_tot = ws.cell(i, 3, f"=IFERROR(INDEX('03_採点シート'!${get_column_letter(TOT)}${FIRST}:${get_column_letter(TOT)}${LAST},"
                           f"MATCH($B{i},'03_採点シート'!${get_column_letter(2)}${FIRST}:${get_column_letter(2)}${LAST},0)),\"\")")
    ai_jdg = ws.cell(i, 4, f"=IFERROR(INDEX('03_採点シート'!${get_column_letter(JDG)}${FIRST}:${get_column_letter(JDG)}${LAST},"
                           f"MATCH($B{i},'03_採点シート'!${get_column_letter(2)}${FIRST}:${get_column_letter(2)}${LAST},0)),\"\")")
    for c in (ai_tot, ai_jdg):
        c.font, c.alignment = BODY, CENTER
    for col in range(1, 5):
        ws.cell(i, col).border = BOX
ws.cell(APP_LAST + 1, 2, "※ 03_採点シートの部門名と1文字違わず同じ表記で入力すること（一致しないとAI参考値が引けない）。").font = SMALL

# ---- ② 役員4名の投票 ----------------------------------------------------
ws.cell(VOTE_HEAD, 2, "② 役員4名の投票（①のAI評価を参考に、上位4部へ順位をつける）").font = H2
head_row(ws, VOTE_HEAD + 1, ["No", "審査員名", "1位", "2位", "3位", "4位", "入力チェック"])
dv_vote = DataValidation(type="list", formula1=f"=${app_rng[1:]}", allow_blank=True,
                          error="①の応募一覧に入力した部門名から選んでください。", errorTitle="入力値エラー")
ws.add_data_validation(dv_vote)
dv_vote.add(f"$C${VOTE_FIRST}:$F${VOTE_LAST}")
for i in range(VOTE_FIRST, VOTE_LAST + 1):
    ws.cell(i, 1, i - VOTE_FIRST + 1).font, ws.cell(i, 1).alignment = BODY, CENTER
    for col in (2, 3, 4, 5, 6):
        ws.cell(i, col).font = INPUT_F
    chk = ws.cell(i, 7,
        f'=IF(COUNTBLANK($C{i}:$F{i})>0,"",'
        f'IF(OR(COUNTIF($C{i}:$F{i},$C{i})>1,COUNTIF($C{i}:$F{i},$D{i})>1,'
        f'COUNTIF($C{i}:$F{i},$E{i})>1,COUNTIF($C{i}:$F{i},$F{i})>1),"★重複あり","OK"))')
    chk.font, chk.alignment = BODY_B, CENTER
    for col in range(1, 8):
        ws.cell(i, col).border = BOX
ws.cell(VOTE_LAST + 1, 2, "※ 1位〜4位は同じ役員の中で重複させないこと（G列が「★重複あり」なら入力し直す）。").font = SMALL

# ---- ③ 得点集計・決勝進出判定 -------------------------------------------
ws.cell(SC_HEAD, 2, "③ 得点集計・決勝進出判定（自動計算）").font = H2
head_row(ws, SC_HEAD + 1,
         ["No", "部門名", "1位票", "2位票", "3位票", "4位票", "合計点\n(40点満点)", "1位保護", "決勝進出", "内部計算用"])
vote_rng = lambda col: f"${col}${VOTE_FIRST}:${col}${VOTE_LAST}"
for k, i in enumerate(range(SC_FIRST, SC_LAST + 1)):
    app_row = APP_FIRST + k
    ws.cell(i, 1, k + 1).font, ws.cell(i, 1).alignment = BODY, CENTER
    b = ws.cell(i, 2, f'=IF($B${app_row}="","",$B${app_row})')
    c1 = ws.cell(i, 3, f'=IF($B{i}="","",COUNTIF({vote_rng("C")},$B{i}))')
    c2 = ws.cell(i, 4, f'=IF($B{i}="","",COUNTIF({vote_rng("D")},$B{i}))')
    c3 = ws.cell(i, 5, f'=IF($B{i}="","",COUNTIF({vote_rng("E")},$B{i}))')
    c4 = ws.cell(i, 6, f'=IF($B{i}="","",COUNTIF({vote_rng("F")},$B{i}))')
    tot = ws.cell(i, 7, f'=IF($B{i}="","",C{i}*10+D{i}*8+E{i}*7+F{i}*6)')
    # 応募4件以下ならこのシート自体を使わない（全件がそのまま決勝進出）ので、
    # ①のメッセージと矛盾しないよう1位保護・決勝進出は空欄にする
    prot = ws.cell(i, 8, f'=IF(OR($B{i}="",COUNTA({app_rng})<=4),"",IF(C{i}>=1,"◎内定",""))')
    key = ws.cell(i, 10, f'=IF($B{i}="","",IF(H{i}<>"",100000,0)+G{i}*100+C{i}*10+D{i})')
    fin = ws.cell(i, 9, f'=IF(OR($B{i}="",COUNTA({app_rng})<=4),"",IF(RANK(J{i},$J${SC_FIRST}:$J${SC_LAST},0)<=4,"◎決勝進出",""))')
    for cell in (b,):
        cell.font, cell.alignment = BODY_B, WRAP
    for cell in (c1, c2, c3, c4, tot):
        cell.font, cell.alignment = BODY, CENTER
    prot.font, prot.alignment = Font(name=FONT, size=10, bold=True, color=NAVY), CENTER
    prot.fill = FILL_GOLD
    fin.font, fin.alignment = Font(name=FONT, size=10, bold=True, color="FFFFFF"), CENTER
    fin.fill = FILL_GREEN
    key.font, key.alignment = SMALL, CENTER
    for col in range(1, 11):
        ws.cell(i, col).border = BOX

lr = SC_LAST + 2
ws.merge_cells(start_row=lr, start_column=2, end_row=lr, end_column=10)
ws.cell(lr, 2,
        "配点：1位10点／2位8点／3位7点／4位6点／5位以下0点（役員4名分の合計・40点満点）。"
        "各役員の1位に選ばれた部門は、合計点に関わらず無条件で決勝進出（1位保護ルール）。"
        "残りの枠は合計点の高い順で埋める。同点の場合は ①1位票の数 → ②2位票の数 → ③4名の合議、で決める。"
        "予選の点数・順位は非公開。本番12名の最終順位には加算しない（登壇順の決定と、本番同点時のタイブレークにのみ使用）。"
        ).font, ws.cell(lr, 2).fill, ws.cell(lr, 2).alignment = SMALL, FILL_YELL, WRAP
ws.row_dimensions[lr].height = 46
ws.merge_cells(start_row=lr + 1, start_column=2, end_row=lr + 1, end_column=10)
ws.cell(lr + 1, 2, "J列（内部計算用）は決勝進出の判定に使う優先度スコアで、触らない。"
                  "1位保護に該当する部門は常に非該当より上位になるよう十分大きい値を足してある。").font = SMALL

ws.freeze_panes = "C6"

# =============================================================== 05 判定ルール
ws = wb.create_sheet("05_判定ルールとフラグ")
banner(ws, "判定バンド・前提フラグ・タイブレーク", 6)
widths = [5, 20, 24, 40, 40, 2]
r = 3
ws.cell(r, 2, "1. 総合判定バンド").font = H2
r += 1
head_row(ws, r, ["", "判定", "総合点", "扱い", ""], widths)
ws.cell(r, 2).value = "判定"
ws.cell(r, 3).value = "総合点"
ws.cell(r, 4).value = "予備審査での扱い"
ws.cell(r, 5).value = ""
r += 1
prev = 101
for cut, lab, desc in BANDS:
    ws.cell(r, 2, lab).font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.cell(r, 3, f"{cut} 点以上" if cut else "50 点未満")
    ws.cell(r, 4, desc)
    for col in (2, 3, 4):
        ws.cell(r, col).border, ws.cell(r, col).alignment = BOX, WRAP
        if col != 2:
            ws.cell(r, col).font = BODY
    ws.row_dimensions[r].height = 24
    r += 1

r += 1
ws.cell(r, 2, "2. 前提フラグ（加点ではなく、該当観点にレベル上限をかける）").font = H2
r += 1
head_row(ws, r, ["", "ID / 名称", "判定の目安", "かかる上限", "根拠となった審査員の言葉"], widths)
ws.cell(r, 2).value = "ID / 名称"
ws.cell(r, 3).value = "どう見分けるか"
ws.cell(r, 4).value = "かかる上限"
ws.cell(r, 5).value = "根拠となった審査員の言葉"
r += 1
for f in FLAGS:
    ws.cell(r, 2, f"{f['id']}  {f['name']}").font = BODY_B
    ws.cell(r, 3, f["detect"]).font = BODY
    cap_txt = " / ".join(
        f"{cid} {next(x['short'] for x in CRITERIA if x['id']==cid)} はレベル {cap} 以下"
        for cid, cap in f["cap"].items())
    ws.cell(r, 4, cap_txt).font = BODY
    ws.cell(r, 5, f["basis"]).font = SMALL
    for col in (2, 3, 4, 5):
        ws.cell(r, col).border, ws.cell(r, col).alignment = BOX, WRAP
    ws.row_dimensions[r].height = 62
    r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.cell(r, 2, "採点シートの W列「上限チェック」が、フラグと入力レベルの矛盾を自動検出する。「上限超過」と出たら、該当観点のレベルを上限まで下げる。").font = SMALL
ws.cell(r, 2).fill, ws.cell(r, 2).alignment = FILL_YELL, WRAP
r += 3

ws.cell(r, 2, "3. 同点時のタイブレーク順").font = H2
r += 1
for i, t in enumerate(TIEBREAK, 1):
    ws.cell(r, 2, f"第{i}順位").font = BODY_B
    ws.cell(r, 2).border, ws.cell(r, 2).fill = BOX, FILL_BLUE
    ws.cell(r, 3, t).font = BODY
    ws.cell(r, 3).border = BOX
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1
ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=5)
ws.cell(r + 1, 2, "10点刻みに丸めた結果、C1・C2 は同じ20点で並ぶ。C6（他部門・全社への波及と未来）は"
                  "単独で最も重い30点なので、まずここのレベル差で決めるのが最も情報量が大きい。"
                  "それでも並ぶ場合はC2→C1の順で見て、最後は事務局・役員の合議とする。").font = SMALL
ws.cell(r + 1, 2).alignment = WRAP

# =============================================================== 05 プロンプト
ws = wb.create_sheet("06_AI審査プロンプト")
banner(ws, "AI予備審査プロンプト（このままコピーして使用）", 3)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 130
ws.column_dimensions["C"].width = 3

from build_prompt import build_prompt  # noqa: E402
prompt_text = build_prompt()

r = 3
ws.cell(r, 2, "B4セル以下がプロンプト本文。セルを選択してコピーし、AIに貼り付けたうえで、応募資料を1件ずつ続けて入力する。").font = SMALL
r += 1
for line in prompt_text.split("\n"):
    c = ws.cell(r, 2, line)
    c.font = Font(name="Consolas", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# =============================================================== 06 元データ
ws = wb.create_sheet("07_元データ_審査員コメント")
banner(ws, "元データ：審査員コメント一覧（全文）", 6, META["source"])
widths = [5, 16, 26, 30, 62, 62]
head_row(ws, 3, ["No", "審査員", "役職", "担当ステージ", "審査コメント（本人の言葉）＝ 従（○）", "グランドフィナーレコメント ＝ 主（◎）"], widths)
ws.row_dimensions[3].height = 34
r = 4
for j in JUDGES:
    vals = [j["no"], j["name"], j["title"], j["stage"], j["stage_c"], j["gf"]]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(r, i, v)
        c.font = BODY_B if i == 2 else BODY
        c.alignment, c.border = WRAP, BOX
        if i == 6:
            c.fill = FILL_GREEN
        elif i == 5:
            c.fill = FILL_YELL
    ws.row_dimensions[r].height = 90
    r += 1
ws.freeze_panes = "C4"

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

out = "AIFES2026_グランドフィナーレ_AI予備審査ルーブリック.xlsx"
wb.save(out)
print("saved", out)
