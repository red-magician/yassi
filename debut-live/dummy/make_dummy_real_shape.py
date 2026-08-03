# 実際のマスターExcel（16_デビュー抽出＋02_採点結果台帳＋04_受賞者マスタが分離しているシート構成）を
# 模したダミーデータ。実データは含まない（氏名・部署はすべて架空）。
import openpyxl, os, random

HERE = os.path.dirname(os.path.abspath(__file__))
random.seed(11)

names = [('Yamada, Taro','AIX本部'), ('Sato, Hanako','製造本部 CS1部'), ('Suzuki, Ichiro','金融保険本部 CS4部'),
          ('Takahashi, Misaki','CTS本部 CSA部'), ('Tanaka, Kenta','事業管理本部'), ('Ito, Sakura','通信サービス本部 CS2部'),
          ('Watanabe, Shota','金融本部 営業部'), ('Nakamura, Yui','CBS事業本部 IPA部'), ('Kobayashi, Akari','HR戦略本部 人事部'),
          ('Kato, Daisuke','PA本部 L&PB推進部'), ('Yoshida, Hina','製造本部 CS2部'), ('Yamamoto, Ren','AIX本部'),
          ('Saito, Yui','事業管理本部'), ('Matsumoto, Yamato','通信サービス本部 CS1部'), ('Inoue, Mizuki','金融本部 営業部')]

wb = openpyxl.Workbook()
ws16 = wb.active
ws16.title = '16_デビュー抽出'
ws16.append(['EntryIndex_RowNo','EntryCode','GlobalEntryId','CompetitionKey','PeriodKey','SubmitterDisplayName',
             'SubmitterEmail','SubmitterUserKey','Department','タイトル','Status','IsHidden',
             'PrimaryFileUrl','ProcedureFileUrl','LikesReceived（もらう）','LikesGiven（押す）'])

rows = []
for i, (name, dept) in enumerate(names, start=1):
    code = f'DEBUT-2026-{i:06d}'
    ukey = f'entra-{i:04d}'
    likes = random.choice([3,8,15,22,28,35,45])       # もらう（25超えは上限テスト用）
    likes_given = random.choice([0,4,13,15,20,34])    # 押す（15超えは上限テスト用・投稿者単位）
    url_ext = 'html' if i != 5 else 'md'   # 1件だけmd
    url = f'https://example.com/DebutSubmissions/entry-{i}.{url_ext}'
    proc = f'https://example.com/DebutSubmissions/entry-{i}-proc.html' if i == 3 else None   # 1件だけ手順書あり
    hidden = 'True' if i == 15 else 'False'   # 1件だけ非公開（除外されるべき）
    rows.append([i, code, f'hackathon-debut-item-{i}', 'debut', 'debut-cool2', name, f'{name.lower()}@example.com',
                 ukey, dept, f'デビュー投稿{i}', 'Published', hidden, url, proc, likes, likes_given])
for r in rows: ws16.append(r)

# 別競技（solo）の行も混ぜて、CompetitionKeyフィルタが効くか確認できるようにする
ws16.append([99, 'SOLO-2026-000099', 'hackathon-solo-item-99', 'solo', 'cool1', 'Other, Person',
             'other@example.com', 'entra-9999', '他本部', 'ソロ投稿', 'Published', 'False',
             'https://example.com/solo-99.html', None, 5, 5])

ws02 = wb.create_sheet('02_採点結果台帳')
ws02.append(['EntryCode','GlobalEntryId','CompetitionKey','PeriodKey','AgentKey（採点主体）',
             '観点①スコア','観点②スコア','観点③スコア','観点④スコア','RubricTotal',
             '提出形式 (HTML/MD)','HTML加点 (+10)','最終スコア','役員エージェント講評'])
# 15件中11件だけAI採点済み（4件は未採点のまま残してテストする）
scored_idx = list(range(1, 12))
for i in scored_idx:
    code = f'DEBUT-2026-{i:06d}'
    axis = [random.choice([2,5,8,10]) for _ in range(4)]
    rubric = round(sum(a/10*25 for a in axis), 1)
    fmt = 'HTML' if i != 5 else 'MD'   # 提出形式は「02_採点結果台帳」側が正で、URL拡張子（i==5だけmd）と一致させる
    ws02.append([code, f'hackathon-debut-item-{i}', 'debut', 'debut-cool2', 'debut-agent',
                 *axis, rubric, fmt, 0, rubric, f'{code} のAI講評ダミー'])
# 他競技の採点行も混ぜる（フィルタで除外されるべき）
ws02.append(['SOLO-2026-000099', 'hackathon-solo-item-99', 'solo', 'cool1', 'solo-agent',
             5,5,5,None, 75, 'HTML', 10, 85, 'ソロのAI講評ダミー'])

ws04 = wb.create_sheet('04_受賞者マスタ')
ws04.append(['SubmitterEntraId','受賞者（本人）','CompetitionKey','PeriodKey','AwardRank','ExcludeFromPeriod','備考'])
# 前月（debut-cool1）にすでに受賞済み、という想定でYamada, Taroを登録（EntraIdでの突合をテストする）
ws04.append(['entra-0001', 'Yamada, Taro', 'debut', 'debut-cool1', 1, '', 'ダミー：前月受賞済み'])

path = os.path.join(HERE, 'master_real_shape.xlsx')
wb.save(path)
print('生成:', path)
print('デビュー対象件数（solo混入・非公開1件を除く）:', len(names))
print('AI採点済み件数:', len(scored_idx))
print('04_受賞者マスタ登録:', 'Yamada, Taro (debut-cool1)')
