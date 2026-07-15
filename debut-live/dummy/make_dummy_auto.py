# 自動集計競技（Buzz/シーンメーカー/クルーアワード）用のダミーマスター。
# 実データの構成（01_EntryINDEX＋03_いいね集計＋04_受賞者マスタ）を模す。氏名・部署はすべて架空。
import openpyxl, os, random

HERE = os.path.dirname(os.path.abspath(__file__))
random.seed(21)

people = [
    ('Aoki, Taro','AIX本部','entra-01'), ('Baba, Hana','製造本部','entra-02'),
    ('Chiba, Ken','金融本部','entra-03'), ('Doi, Mio','人事部','entra-04'),
    ('Endo, Sho','営業1部','entra-05'), ('Fujita, Rin','開発2部','entra-06'),
    ('Goto, Yui','経理部','entra-07'), ('Hara, Dai','企画部','entra-08'),
    ('Ito, Aoi','CTS本部','entra-09'), ('Kato, Rei','CBS本部','entra-10'),
    ('Mori, Nao','通信本部','entra-11'), ('Sato, Jun','PA本部','entra-12'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '01_EntryINDEX'
HEAD = ['EntryCode','CompetitionKey','PeriodKey','SubmitterDisplayName','SubmitterEmail','SubmitterEntraId',
        'Department','IsBuzzEligible','IsSceneEligible','SupportAmbassadorNames','審査対象フラグ']
ws.append(HEAD)

likes_rows = []  # (EntryCode, LikesReceived)
eid = 0
def add(comp, name, dept, entra, likes, buzz=False, scene=False, amb='', target=True):
    global eid; eid += 1
    code = f'{comp.upper()}-2026-{eid:06d}'
    ws.append([code, comp, f'{comp}-cool1', name, f'{name.split(",")[0].lower()}@example.com', entra,
               dept, 'TRUE' if buzz else 'False', 'TRUE' if scene else 'False', amb, 'TRUE' if target else 'False'])
    likes_rows.append((code, likes))
    return code

# --- Buzz: 5人が複数投稿。平均タイブレークが効くケースを仕込む ---
# Aoki: 3投稿で合計60（平均20） / Baba: 1投稿で55（平均55） / Chiba: 10投稿で合計50（平均5）
for lk in [25,20,15]: add('buzz', *people[0], lk, buzz=True)
add('buzz', *people[1], 55, buzz=True)
for _ in range(10): add('buzz', *people[2], 5, buzz=True)
add('buzz', *people[3], 40, buzz=True)
add('buzz', *people[4], 33, buzz=True)
add('buzz', *people[5], 18, buzz=True)
add('buzz', *people[6], 12, buzz=True)
add('buzz', *people[7], 9, buzz=True)   # 8人 → 上位6名選出可

# --- Scene: 対象者は scene-eligible フラグ。likesは全競技横断で集計されるので、各人の他競技投稿にlikesを付ける ---
# scene-eligible: Sato, Mori, Kato, Ito, Hara, Goto の6人（6名ちょうど）
scene_people = people[6:12]  # Goto..Sato
for pn in scene_people:
    add('scene', *pn, 0, scene=True)          # scene投稿自体はlikes0（実データと同じ傾向）
# 彼らのsolo投稿に likes を付けて通期累計に反映
solo_likes = {6:30, 7:12, 8:48, 9:5, 10:22, 11:40}  # index→likes
for i,pn in zip(range(6,12), scene_people):
    add('solo', *pn, solo_likes[i])

# --- Crew: 支援アンバサダー名を色々な表記で記載（名寄せテスト） ---
# 「城間 康」を色々な表記で。複数名セルも。
crew_ambs = [
    '城間 康',
    '城間 康\n大坪 敏雄',   # 複数名（改行）
    '城間さん',              # 表記ゆれ → 城間 康 にマージすべき
    '大坪 敏雄',
    '色部 晟洋',
    '色部さん',              # 表記ゆれ → 色部 晟洋
    'ぺぺ',                  # あだ名
    'ぺぺ',
    'くぁwせdrftgy',        # ゴミ → 無効指定すべき
    '城間 康, 色部 晟洋',    # カンマ複数名
]
for i, amb in enumerate(crew_ambs):
    # 支援者名はどのエントリーにも付きうる（応募者は適当）
    add(random.choice(['solo','collab','jam','debut']), *people[i%12], random.choice([2,5,8]), amb=amb)

# 03_いいね集計
ws2 = wb.create_sheet('03_いいね集計')
ws2.append(['EntryCode','CompetitionKey','SubmitterDisplayName','LikesReceived（もらう）','自己リアクション行数'])
# EntryCode→comp/name lookup from ws
idx_rows = list(ws.iter_rows(min_row=2, values_only=True))
meta = {r[0]:(r[1], r[3]) for r in idx_rows}
for code, lk in likes_rows:
    comp, name = meta[code]
    ws2.append([code, comp, name, lk, 0])

# 04_受賞者マスタ（Buzz中間で Endo が受賞済み → 最終で除外される想定）
ws3 = wb.create_sheet('04_受賞者マスタ')
ws3.append(['SubmitterEntraId','受賞者（本人）','CompetitionKey','PeriodKey','AwardRank','ExcludeFromPeriod','備考'])
ws3.append(['entra-05','Endo, Sho','buzz','buzz-cool1',1,'','ダミー：Buzz中間受賞済み'])

path = os.path.join(HERE, 'master_auto.xlsx')
wb.save(path)
print('生成:', path)
print(' Buzz対象:8人 / Scene対象:6人 / Crew支援名:10記載')
print(' 04_受賞者マスタ: Endo, Sho（buzz）を登録 → Buzz最終で除外されるはず')
