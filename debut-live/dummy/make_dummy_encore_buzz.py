from openpyxl import Workbook

wb = Workbook()
ws0 = wb.active
ws0.title = "README"
ws0.append(["項目", "内容"])
ws0.append(["ダミーデータ", "テスト用"])

head = ['集計区分','EntryCode','タイトル','GlobalEntryId','CompetitionDisplayNameSnapshot','CompetitionKey',
        'PeriodKey','Status','SubmissionStatus','IsHidden','IsEncoreEligible','PublishedAt',
        'SubmitterDisplayName','Department','LikeCount','UniqueUserCount','ReactionRows',
        'PrimaryFileUrl','TeamsPostUrl']

ws1 = wb.create_sheet("Encore_by_entry")
ws1.append(head)
encore_data = [
    ("DEBUT-2026-0001", "議事録を10分で仕上げる話", "Tanaka, Kenta", "開発1部", 92),
    ("SOLO-2026-0002", "工数分析から解放された話", "Sato, Yui", "営業部", 80),
    ("DEBUT-2026-0003", "AIチャットで一次対応を自動化", "Suzuki, Ren", "CS部", 75),
    ("SOLO-2026-0004", "提案書ドラフト自動生成", "Takahashi, Mio", "企画部", 70),
    ("DEBUT-2026-0005", "校正アシスタントを作った話", "Ito, Sora", "品質保証部", 65),
    ("DEBUT-2026-0006", "会議アジェンダ自動整形", "Watanabe, Rio", "総務部", 60),
    ("SOLO-2026-0007", "日報要約フローを構築", "Yamamoto, Kai", "人事部", 55),
    ("DEBUT-2026-0008", "経費精算チェックを自動化", "Nakamura, Yui", "経理部", 50),
]
for code, title, name, dept, likes in encore_data:
    ws1.append(["アンコール", code, title, f"gid-{code}", "デビューライブ", "debut", "debut-cool1",
                "Published", "Published", "FALSE", "TRUE", "2026/7/1 10:00",
                name, dept, likes, likes, likes,
                f"https://example.com/works/{code}.html", ""])

ws2 = wb.create_sheet("Buzz_by_entry")
ws2.append(head)
buzz_data = [
    ("BUZZ-2026-0001", "AI駆動開発とは？を解説", "Kobayashi, Sho", "AIX本部", 44),
    ("BUZZ-2026-0002", "Copilot Studioのモデル選択", "Kato, Nana", "CMS本部", 38),
    ("BUZZ-2026-0003", "生成AI従量課金時代の使い所", "Yoshida, Ken", "HR本部", 35),
    ("BUZZ-2026-0004", "隠れた神機能Notebook", "Yamada, Aoi", "AIX本部", 29),
    ("BUZZ-2026-0005", "New experienceを試した", "Saito, Riku", "HR本部", 24),
    ("BUZZ-2026-0006", "後でやろうで忘れないコツ", "Hasegawa, Emi", "製造本部", 19),
    ("BUZZ-2026-0007", "予約してCopilotを働かせる", "Fujii, Taku", "AIX本部", 15),
]
for code, title, name, dept, likes in buzz_data:
    ws2.append(["バズ", code, title, f"gid-{code}", "バズステージ", "buzz", "buzz-cool1",
                "Published", "Published", "FALSE", "TRUE", "2026/7/1 10:00",
                name, dept, likes, likes, likes,
                f"https://example.com/works/{code}.html", ""])

wb.save("master_encore_buzz.xlsx")
print("done")
