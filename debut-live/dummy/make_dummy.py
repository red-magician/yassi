import openpyxl, os, random, shutil

HERE = os.path.dirname(os.path.abspath(__file__))
random.seed(7)

HEAD = ['EntryCode','GlobalEntryId','対象月','応募者・部署','作品リンク','ProcedureFileUrl',
        '観点①スコア','観点②スコア','観点③スコア','観点④スコア',
        '作品ファイル形式','手順書ファイル形式','いいね（もらう）','AI講評']

names = ['山田太郎','佐藤花子','鈴木一郎','高橋美咲','田中健太','伊藤さくら','渡辺翔太','中村optimoptim',
         '小林あかり','加藤大輔','吉田陽菜','山本蓮','斎藤結衣','松本大和','井上美月']
depts = ['営業1課','開発2課','企画部','人事部','経理部']
fmts = ['HTML','PDF','Word','HTML','PowerPoint','HTML']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '02_採点結果台帳_デビューライブ'
ws.append(HEAD)

for i, name in enumerate(names, start=1):
    code = f'DEBUT-2026-{i:04d}'
    who = f'{name}/{random.choice(depts)}'
    axis = [random.choice([2,5,8,10]) for _ in range(4)]
    fmt_work = random.choice(fmts)
    fmt_proc = random.choice(fmts)
    likes = random.choice([2,5,10,18,25,30,40])   # 25超えは上限テスト用
    has_work_link = i != 12   # 1件だけ資料欠損にする
    row = [code, f'GID-{i:05d}', '2026年7月', who,
           (f'https://example.com/works/{code}' if has_work_link else ''),
           f'https://example.com/proc/{code}',
           *axis, fmt_work, fmt_proc, likes, f'{name}さんの体験談（AI生成ダミー講評）']
    ws.append(row)

master_jp = os.path.join(HERE, 'ダミー_採点マスター_デビューライブ_7月.xlsx')
wb.save(master_jp)

# ---- 前月(6月)の受賞一覧（除外リストのテスト用。山田太郎・佐藤花子が6月に受賞済みという設定）----
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = '受賞一覧'
ws2.append(['対象月','受賞順位','EntryCode','応募者・部署','作品リンク','ルーブリック合計','いいね（もらう）','いいね加点','HTML加点','最終スコア','受賞理由','講評コメント'])
ws2.append(['2026年6月', 1, 'DEBUT-2026-0090', '山田太郎/営業1課', 'https://example.com/works/DEBUT-2026-0090', 90, 20, 20, 10, 120, '見本の受賞理由', '見本の講評'])
ws2.append(['2026年6月', 2, 'DEBUT-2026-0091', '佐藤花子/開発2課', 'https://example.com/works/DEBUT-2026-0091', 85, 15, 15, 5, 105, '見本の受賞理由2', '見本の講評2'])
past_jp = os.path.join(HERE, 'ダミー_受賞一覧_6月.xlsx')
wb2.save(past_jp)

# ---- headless Chromiumの非ASCIIファイル名バグを避けるためのASCII名コピー（自動テスト用フィクスチャ）----
master_ascii = os.path.join(HERE, 'master_test.xlsx')
past_ascii = os.path.join(HERE, 'past_test.xlsx')
shutil.copy(master_jp, master_ascii)
shutil.copy(past_jp, past_ascii)

# ---- 突合テスト用：DEBUT-2026-0008（受賞候補圏外）の観点①だけ変えたマスター（スコア不一致検知の確認用）----
wb3 = openpyxl.load_workbook(master_ascii)
ws3 = wb3.active
head3 = [c.value for c in next(ws3.iter_rows(min_row=1, max_row=1))]
code_col = head3.index('EntryCode') + 1
ax1_col = head3.index('観点①スコア') + 1
for row in ws3.iter_rows(min_row=2):
    if row[code_col-1].value == 'DEBUT-2026-0008':
        row[ax1_col-1].value = 2
staffb_ascii = os.path.join(HERE, 'master_test_staffB.xlsx')
wb3.save(staffb_ascii)

print('生成完了:')
for p in (master_jp, past_jp, master_ascii, past_ascii, staffb_ascii):
    print(' -', p)
