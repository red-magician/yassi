#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""WBS(城間さん担当) → Planner取り込み用に整形して出力する。"""
import csv, datetime, re, os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import sys
# 使い方: python3 wbs_to_planner.py <WBS.xlsx> [担当者名]
SRC = sys.argv[1] if len(sys.argv) > 1 else "AIFES2026_WBS.xlsx"
OUT = os.path.dirname(os.path.abspath(__file__))
os.makedirs(OUT, exist_ok=True)
YEAR = 2026
ME = sys.argv[2] if len(sys.argv) > 2 else "城間"

wb = load_workbook(SRC, data_only=True)
ws = wb["WBS"]
rows = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]

# --- 親子判定 ---
parent_name, kind = None, {}
kids = {}
for r in rows:
    t = str(r[3] or "")
    if t.startswith("└"):
        kind[r[0]] = ("child", parent_name)
        kids.setdefault(parent_name, []).append(r)
    else:
        parent_name = t
        kind[r[0]] = ("root", None)

# 親の短縮ラベル（子タスク名の接頭辞にする）
def short(p):
    p = re.sub(r"（[^）]*）", "", p or "").strip()
    p = p.replace("集計・受賞者決定", "集計").replace("事前審査〜決勝進出4部の決定", "事前審査")
    p = p.replace("審査シート配布・回収", "審査シート").replace(" AI審査", " AI審査")
    return p[:18]

def to_date(s):
    s = str(s or "").strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if not m:
        return ""
    return f"{YEAR}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

PRI = {"高": "重要", "中": "中", "低": "低"}
BUCKET = {
    "1. 共通基盤・システム改修": "1_共通基盤",
    "2. 第2クール（8/14締切）": "2_第2クール",
    "3. 役員審査ステージ 第2クール": "3_役員審査",
    "4. グランドフィナーレ": "4_GF",
    "5. 第3クール／FINALE（9/9締切）": "5_第3クール",
    "6. 広報・アンバサダー": "6_広報",
    "8. 振り返り・クロージング": "8_振り返り",
}

hits = [r for r in rows if r[4] and ME in str(r[4])]
summary_rows, tasks = [], []

for r in hits:
    no, phase, mid, task, tantou, kosu, start, due, pri, status, note = r[:11]
    k, p = kind[no]
    task = str(task or "").strip()
    # 子を持つまとめ行は除外（工数二重計上を避ける）→ 参考リストへ
    if k == "root" and task in kids:
        summary_rows.append(dict(no=no, task=task, kosu=kosu, due=to_date(due),
                                 n_kids=len(kids[task]),
                                 n_mine=len([c for c in kids[task] if c[4] and ME in str(c[4])])))
        continue
    if k == "child":
        title = f"【{short(p)}】{task.lstrip('└ ').strip()}"
    else:
        title = task
    tantou = str(tantou or "").strip()
    solo = (tantou == ME) or tantou.startswith(f"{ME} ←")
    tasks.append(dict(
        no=no,
        bucket=BUCKET.get(str(phase).strip(), str(phase)),
        title=title,
        start=to_date(start),
        due=to_date(due),
        pri=PRI.get(str(pri).strip(), "中"),
        kosu=("" if kosu in (None, "") else kosu),
        tantou=tantou,
        solo="単独" if solo else "共同",
        status=str(status or "").strip(),
        note=str(note or "").strip(),
        mid=str(mid or "").strip(),
    ))

tasks.sort(key=lambda t: (t["due"] or "9999", t["no"]))

HEAD = ["タスク名", "バケット", "開始日", "期限", "優先度", "進捗", "メモ"]
def memo(t):
    bits = [f"WBS No.{t['no']}", f"担当:{t['tantou']}({t['solo']})"]
    if t["kosu"]:
        bits.append(f"工数{t['kosu']}人日")
    if t["status"] and t["status"] not in ("未着手",):
        bits.append(f"WBS状態:{t['status']}")
    if t["note"]:
        bits.append(t["note"])
    return " / ".join(bits)

def progress(t):
    s = t["status"]
    return {"完了": "完了", "準備中": "進行中", "調査中": "進行中"}.get(s, "未開始")

recs = [[t["title"], t["bucket"], t["start"], t["due"], t["pri"], progress(t), memo(t)] for t in tasks]

# CSV (Excel向けBOM付き)
with open(f"{OUT}/{ME}さん_Planner取込.csv", "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f); w.writerow(HEAD); w.writerows(recs)
# TSV (グリッド貼り付け用)
with open(f"{OUT}/{ME}さん_Planner取込.tsv", "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f, delimiter="\t"); w.writerow(HEAD); w.writerows(recs)
# タスク名のみ（Planner一括追加用）
with open(f"{OUT}/{ME}さん_タスク名のみ.txt", "w", encoding="utf-8") as f:
    for b in sorted({t["bucket"] for t in tasks}):
        grp = [t for t in tasks if t["bucket"] == b]
        f.write(f"===== バケット: {b} （{len(grp)}件） =====\n")
        for t in grp:
            f.write(t["title"] + "\n")
        f.write("\n")

# XLSX
out = Workbook(); s = out.active; s.title = "Planner取込"
s.append(HEAD)
for c in s[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3A5F")
    c.alignment = Alignment(horizontal="center")
for r in recs:
    s.append(r)
for i, w_ in enumerate([46, 14, 12, 12, 9, 9, 60], 1):
    s.column_dimensions[chr(64 + i)].width = w_
s2 = out.create_sheet("参考_まとめ行(登録対象外)")
s2.append(["WBS No.", "まとめタスク", "工数(人日)", "全体期限", "子タスク数", "うち城間"])
for c in s2[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="8A5A2B")
for m in summary_rows:
    s2.append([m["no"], m["task"], m["kosu"], m["due"], m["n_kids"], m["n_mine"]])
for i, w_ in enumerate([10, 46, 12, 14, 12, 10], 1):
    s2.column_dimensions[chr(64 + i)].width = w_
out.save(f"{OUT}/{ME}さん_Planner取込.xlsx")

print(f"登録対象タスク: {len(tasks)}件 / まとめ行(除外): {len(summary_rows)}件")
from collections import Counter
for b, n in sorted(Counter(t["bucket"] for t in tasks).items()):
    print(f"  {b}: {n}件")
print(f"\n単独担当: {sum(1 for t in tasks if t['solo']=='単独')}件 / 共同: {sum(1 for t in tasks if t['solo']=='共同')}件")
