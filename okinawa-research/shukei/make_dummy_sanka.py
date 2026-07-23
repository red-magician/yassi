#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
拠点別参加率の集計（sanka_ritsu.py）の動作確認用ダミーマスターを生成する。

実データのマスターExcelに寄せて、次の3シートを持つブックを作る:
  01_EntryINDEX : エントリー1行=1投稿。EntryCode / CompetitionKey / PeriodKey /
                  SubmitterDisplayName / SubmitterEntraId / Department / 拠点
  03_いいね集計  : 参考（集計には未使用。実データ構成に合わせて同梱するだけ）
  00_社員名簿    : 拠点ごとの在籍人数（エントリー率の分母）。列= 拠点 / 在籍人数

沖縄が「エントリー率（入口）」も「アクティブ率（継続=複数投稿者の割合）」も
高くなるように作ってある（＝集計スクリプトが沖縄の突出を検出できることの確認用）。

出力: ./dummy_master_sanka.xlsx
実行: python3 make_dummy_sanka.py
"""
import random
from openpyxl import Workbook

random.seed(20260723)  # 再現性のため固定

# 拠点: (在籍人数, 参加者率の目安, 複数投稿者になりやすさ)
# 沖縄を高めに設定
OFFICES = {
    "沖縄":   dict(headcount=40,  entry_rate=0.65, repeat=0.70),
    "東京":   dict(headcount=180, entry_rate=0.22, repeat=0.35),
    "大阪":   dict(headcount=90,  entry_rate=0.28, repeat=0.30),
    "名古屋": dict(headcount=60,  entry_rate=0.20, repeat=0.25),
    "福岡":   dict(headcount=50,  entry_rate=0.30, repeat=0.32),
    "札幌":   dict(headcount=35,  entry_rate=0.17, repeat=0.20),
}
DEPTS = ["営業部", "開発部", "管理部", "CS部", "企画部"]
COMPS = ["debut", "buzz", "scene", "crew"]
PERIOD = "2026-07"

wb = Workbook()

# --- 01_EntryINDEX ---
ws = wb.active
ws.title = "01_EntryINDEX"
ws.append(["EntryCode", "CompetitionKey", "PeriodKey",
           "SubmitterDisplayName", "SubmitterEntraId", "Department", "拠点"])

entry_no = 0
for office, cfg in OFFICES.items():
    hc = cfg["headcount"]
    n_entrants = max(1, round(hc * cfg["entry_rate"]))
    for i in range(n_entrants):
        pid = f"{office[:2]}{i:03d}"
        entra = f"{office}-{i:03d}@example.local"     # 人の同定キー(EntraId相当)
        name = f"{office}メンバー{i+1:02d}"
        dept = random.choice(DEPTS)
        # この人の投稿数: repeat の確率で複数投稿(=アクティブ)
        if random.random() < cfg["repeat"]:
            posts = random.randint(2, 5)
        else:
            posts = 1
        for _ in range(posts):
            entry_no += 1
            comp = random.choice(COMPS)
            ws.append([f"E-2026-{entry_no:04d}", comp, PERIOD,
                       name, entra, dept, office])

# --- 03_いいね集計（参考同梱・集計には未使用） ---
ws2 = wb.create_sheet("03_いいね集計")
ws2.append(["EntryCode", "SubmitterDisplayName", "LikesReceived（もらう）"])
for row in ws.iter_rows(min_row=2, values_only=True):
    ws2.append([row[0], row[3], random.randint(0, 40)])

# --- 00_社員名簿（エントリー率の分母） ---
ws3 = wb.create_sheet("00_社員名簿")
ws3.append(["拠点", "在籍人数"])
for office, cfg in OFFICES.items():
    ws3.append([office, cfg["headcount"]])

out = "dummy_master_sanka.xlsx"
wb.save(out)

total_posts = entry_no
print(f"生成: {out}")
print(f"  01_EntryINDEX : {total_posts}投稿")
print(f"  00_社員名簿    : {len(OFFICES)}拠点")
print(f"  期待される傾向 : 沖縄がエントリー率・アクティブ率とも最上位")
