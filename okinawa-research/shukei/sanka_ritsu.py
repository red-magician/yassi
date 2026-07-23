#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
拠点別 参加率 集計スクリプト  ―  AI FES 2026 測定班 / 沖縄横展開

集計マスターExcelを読み込み、拠点ごとに次の2指標を出す:
  ・エントリー率 = ユニーク参加者数 ÷ 拠点の在籍人数      （入口の広さ）
  ・アクティブ率 = 複数回投稿した人の数 ÷ ユニーク参加者数  （継続の強さ）
沖縄が全社平均に対してどれだけ突出しているかを可視化する。

■ 使い方
    python3 sanka_ritsu.py <マスター.xlsx> [--out 出力.xlsx] [--highlight 沖縄]

  実データが用意できないうちは、同フォルダの make_dummy_sanka.py で
  ダミーを作って動作確認できる:
    python3 make_dummy_sanka.py
    python3 sanka_ritsu.py dummy_master_sanka.xlsx

■ 読み込む列（別名を自動で許容）
  参加(エントリー)シート … シート名に "EntryINDEX" を含む、または EntryCode 列を持つ最初のシート
    - EntryCode                （必須：投稿1件を表す）
    - 拠点                     （'拠点' / '事業所' / '勤務地' / 'Office' / 'Location' / '拠点名'）
    - 人の同定キー              （'SubmitterEntraId' > 'SubmitterEmail' > 'SubmitterUserKey' > 'SubmitterDisplayName'）
  在籍人数シート（分母・任意） … シート名に "社員" or "名簿" or "roster" を含む
    - 拠点  と  在籍人数（'在籍人数' / '人数' / '社員数' / 'headcount' / 'Headcount'）
    ※ 無い場合はエントリー率を "-"（分母不明）とし、参加者数とアクティブ率のみ出す。

集計対象は全競技（debut/buzz/scene/crew 等）を合算した「人の参加」。
"""
import sys
import argparse
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- 列名の候補（実データの表記ゆれを吸収） ----
OFFICE_COLS = ["拠点", "事業所", "勤務地", "拠点名", "Office", "office", "Location", "location"]
PERSON_COLS = ["SubmitterEntraId", "SubmitterEmail", "SubmitterUserKey", "SubmitterDisplayName", "応募者・部署"]
ENTRY_CODE_COLS = ["EntryCode", "エントリーコード"]
HEADCOUNT_COLS = ["在籍人数", "人数", "社員数", "在籍数", "headcount", "Headcount"]


def _norm(v):
    return str(v).strip() if v is not None else ""


def _pick_col(header, candidates):
    """ヘッダー行(list)から候補名に一致する列インデックスを返す。無ければ None。"""
    idx = {(_norm(h)): i for i, h in enumerate(header)}
    for c in candidates:
        if c in idx:
            return idx[c]
    return None


def _sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


def find_entry_sheet(wb):
    # 1) シート名に EntryINDEX を含む
    for ws in wb.worksheets:
        if "entryindex" in ws.title.lower().replace("_", ""):
            return ws
    # 2) EntryCode 列を持つ最初のシート
    for ws in wb.worksheets:
        header, _ = _sheet_rows(ws)
        if _pick_col(header, ENTRY_CODE_COLS) is not None:
            return ws
    return None


def find_roster_sheet(wb):
    for ws in wb.worksheets:
        t = ws.title.lower()
        if ("社員" in ws.title) or ("名簿" in ws.title) or ("roster" in t) or ("在籍" in ws.title):
            header, _ = _sheet_rows(ws)
            if _pick_col(header, OFFICE_COLS) is not None and _pick_col(header, HEADCOUNT_COLS) is not None:
                return ws
    return None


def load_headcounts(wb):
    ws = find_roster_sheet(wb)
    if ws is None:
        return None
    header, rows = _sheet_rows(ws)
    oi = _pick_col(header, OFFICE_COLS)
    hi = _pick_col(header, HEADCOUNT_COLS)
    hc = {}
    for r in rows:
        office = _norm(r[oi]) if oi is not None and oi < len(r) else ""
        if not office:
            continue
        try:
            hc[office] = int(float(r[hi]))
        except (TypeError, ValueError):
            pass
    return hc or None


def aggregate(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    ews = find_entry_sheet(wb)
    if ews is None:
        sys.exit("エラー: EntryCode を持つ参加シートが見つかりません（01_EntryINDEX 等）。")
    header, rows = _sheet_rows(ews)

    oi = _pick_col(header, OFFICE_COLS)
    if oi is None:
        sys.exit("エラー: 拠点を示す列が見つかりません（" + " / ".join(OFFICE_COLS) + "）。\n"
                 "→ 実データに拠点/事業所/勤務地の列を足してから再実行してください。")
    pi = _pick_col(header, PERSON_COLS)
    if pi is None:
        sys.exit("エラー: 人の同定キー列が見つかりません（" + " / ".join(PERSON_COLS) + "）。")

    # 拠点 -> 人 -> 投稿数
    posts = defaultdict(lambda: defaultdict(int))
    for r in rows:
        if oi >= len(r) or pi >= len(r):
            continue
        office = _norm(r[oi])
        person = _norm(r[pi]).lower()  # メール・表示名の表記ゆれを軽く吸収
        if not office or not person:
            continue
        posts[office][person] += 1

    headcounts = load_headcounts(wb)

    results = []
    for office, people in posts.items():
        entrants = len(people)                                  # ユニーク参加者
        active = sum(1 for n in people.values() if n >= 2)      # 複数投稿=継続
        total_posts = sum(people.values())
        hc = headcounts.get(office) if headcounts else None
        entry_rate = (entrants / hc) if hc else None
        active_rate = (active / entrants) if entrants else 0.0
        results.append(dict(
            office=office, headcount=hc, entrants=entrants, active=active,
            posts=total_posts, entry_rate=entry_rate, active_rate=active_rate,
        ))

    # エントリー率（無ければ参加者数）の降順
    results.sort(key=lambda d: (d["entry_rate"] is not None, d["entry_rate"] or 0, d["entrants"]), reverse=True)
    return results, (headcounts is not None)


def summarize(results):
    """全社合計・平均の行を作る。"""
    tot_hc = sum(r["headcount"] for r in results if r["headcount"]) or None
    tot_entrants = sum(r["entrants"] for r in results)
    tot_active = sum(r["active"] for r in results)
    tot_posts = sum(r["posts"] for r in results)
    return dict(
        office="全社", headcount=tot_hc, entrants=tot_entrants, active=tot_active,
        posts=tot_posts,
        entry_rate=(tot_entrants / tot_hc) if tot_hc else None,
        active_rate=(tot_active / tot_entrants) if tot_entrants else 0.0,
    )


def _pct(x):
    return f"{x*100:.1f}%" if x is not None else "-"


def print_table(results, total, highlight, has_hc):
    print("\n=== 拠点別 参加率 ===")
    if not has_hc:
        print("（在籍人数シートが無いため エントリー率は '-'。参加者数とアクティブ率のみ）")
    head = f"{'拠点':<8}{'在籍':>6}{'参加者':>7}{'エントリー率':>12}{'継続者':>7}{'アクティブ率':>12}{'投稿':>6}"
    print(head)
    print("-" * len(head))
    for r in results:
        mark = " ★" if r["office"] == highlight else ""
        print(f"{r['office']:<8}{(r['headcount'] or '-'):>6}{r['entrants']:>7}"
              f"{_pct(r['entry_rate']):>12}{r['active']:>7}{_pct(r['active_rate']):>12}{r['posts']:>6}{mark}")
    print("-" * len(head))
    print(f"{total['office']:<8}{(total['headcount'] or '-'):>6}{total['entrants']:>7}"
          f"{_pct(total['entry_rate']):>12}{total['active']:>7}{_pct(total['active_rate']):>12}{total['posts']:>6}")

    # 沖縄の突出度
    hl = next((r for r in results if r["office"] == highlight), None)
    if hl:
        print(f"\n--- {highlight} vs 全社平均 ---")
        if hl["entry_rate"] and total["entry_rate"]:
            ratio = hl["entry_rate"] / total["entry_rate"]
            print(f"エントリー率: {highlight} {_pct(hl['entry_rate'])}  /  全社 {_pct(total['entry_rate'])}  → 全社比 {ratio:.2f}倍")
        if total["active_rate"]:
            ratio = hl["active_rate"] / total["active_rate"] if total["active_rate"] else 0
            print(f"アクティブ率: {highlight} {_pct(hl['active_rate'])}  /  全社 {_pct(total['active_rate'])}  → 全社比 {ratio:.2f}倍")
        print("読み解き: エントリー率が高い=入口が広い / アクティブ率が高い=続く仕組みがある。"
              " どちらが強みかで横展開の打ち手が変わる。")


def write_excel(results, total, highlight, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "拠点別参加率"

    navy = "1F3A5F"; gold = "C8A24B"
    thin = Side(style="thin", color="D8DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["拠点", "在籍人数", "参加者数", "エントリー率", "継続者数(複数投稿)", "アクティブ率", "総投稿数"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    def _row(r, bold=False, fill=None):
        ws.append([
            r["office"], r["headcount"] if r["headcount"] else "-", r["entrants"],
            round(r["entry_rate"], 4) if r["entry_rate"] is not None else "-",
            r["active"],
            round(r["active_rate"], 4) if r["active_rate"] is not None else "-",
            r["posts"],
        ])
        row = ws[ws.max_row]
        for c in row:
            c.border = border
            if bold:
                c.font = Font(bold=True)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        # 率をパーセント表示に
        for i in (4, 6):
            cell = row[i - 1]
            if isinstance(cell.value, float):
                cell.number_format = "0.0%"

    for r in results:
        _row(r, fill="FBF3DD" if r["office"] == highlight else None)
    _row(total, bold=True, fill="EEF2F7")

    widths = [12, 10, 10, 14, 18, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.append([])
    note = ws.cell(row=ws.max_row + 1, column=1,
                   value="エントリー率=参加者数÷在籍人数（入口の広さ） / アクティブ率=複数投稿者÷参加者数（継続の強さ）")
    note.font = Font(size=9, italic=True, color="5B6B7A")

    wb.save(out_path)
    print(f"\nExcel出力: {out_path}")


def main():
    ap = argparse.ArgumentParser(description="拠点別 参加率 集計")
    ap.add_argument("master", help="集計マスターExcelのパス")
    ap.add_argument("--out", default="拠点別参加率.xlsx", help="出力Excel（既定: 拠点別参加率.xlsx）")
    ap.add_argument("--highlight", default="沖縄", help="強調する拠点名（既定: 沖縄）")
    args = ap.parse_args()

    results, has_hc = aggregate(args.master)
    total = summarize(results)
    print_table(results, total, args.highlight, has_hc)
    write_excel(results, total, args.highlight, args.out)


if __name__ == "__main__":
    main()
